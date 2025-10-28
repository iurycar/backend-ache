from string import Template
from pathlib import Path
from .db import ENGINE
import pandas as pd
import openpyxl

diretorio = Path(__file__).parent
path = diretorio / "uploads" 

def to_mysql_inserir(id_file: str) -> None:
    try:
        file_path = path / id_file

        # Verifica se o arquivo existe antes de tentar inserir
        if not file_path.exists():
            print(f"Erro: O arquivo {file_path} não foi encontrado.")
            return

        sheet: pd.DataFrame = pd.read_excel(file_path)

        # Adiciona o id_file como uma coluna
        sheet['id_file'] = id_file

        # Renomeia as colunas para corresponder ao banco de dados
        sheet.rename(columns={
                'Classificação': 'classe',
                'Categoria': 'category',
                'Fase': 'phase',
                'Condição': 'status',
                'Nome': 'name',
                'Duração': 'duration',
                'Como Fazer': 'text',
                'Documento Referência': 'reference',
                '% Concluída': 'conclusion'
            }, inplace=True)

        if 'Número' in sheet.columns:
            sheet.rename(columns={'Número': 'num'}, inplace=True)
        else:
            # Se não tiver a coluna 'Número', cria uma nova coluna com valores sequenciais
            sheet.insert(0, 'num', range(1, len(sheet) + 1))
        
        # Extrai as URLs da coluna documento referência e atualiza a coluna 'reference'
        # Pois o pandas não consegue extrair hyperlinks diretamente do Excel
        if 'reference' in sheet.columns:
            urls = extract_reference_urls(file_path, cabecalho='Documento Referência')
            
            if urls is not None and len(urls) > 0:
                urls = urls[:len(sheet)] # Garante que o número de URLs não exceda o número de linhas no DataFrame

                # Pega cada índice do DataFrame e atualiza a coluna 'reference' com a URL correspondente
                for i, index in enumerate(sheet.index):
                    if i < len(urls) and urls[i]:
                        sheet.at[index, 'reference'] = urls[i] # Atualiza a célula na linha 'index' e coluna 'reference' com a URL extraída
                    else:
                        if pd.notna(sheet.at[index, 'reference']):
                            sheet.at[index, 'reference'] = sheet.at[index, 'reference']

            sheet['reference'] = sheet['reference'].fillna('').astype(str)

        print("Iniciando a inserção na tabela SHEET...")
        sheet.to_sql(
            'SHEETS',
            con=ENGINE,
            if_exists='append',
            index=False
        )
        
        print(f"Dados inseridos com sucesso na tabela SHEETS.")
    except Exception as error:
        print(f"Erro ao inserir os dados do arquivo {id_file} para o MySQL: {error}")


def from_mysql_extrair(id_file:str) -> None:
    try:
        file_path = path / id_file

        colunas = {
            'num': 'Número',
            'classe': 'Classificação',
            'category': 'Categoria',
            'phase': 'Fase',
            'status': 'Condição',
            'name': 'Nome',
            'duration': 'Duração',
            'text': 'Como Fazer',
            'reference': 'Documento Referência',
            'conclusion': '% Concluída',
        }

        # Consulta SQL para buscar os dados na tabela SHEETS com base no id_file
        template_sql_query = Template("SELECT `num`,`classe`,`category`,`phase`,`status`,`name`,CONCAT(`duration`, ' dias') AS `duration`, `text`,`reference`,`conclusion` FROM `SHEETS` WHERE `id_file` = '$id_file'")
        sql_query = template_sql_query.safe_substitute(id_file=id_file)

        print(f"Buscando dados na tabela SHEETS para o arquivo '{id_file}'.")

        sheet: pd.DataFrame = pd.read_sql(sql_query, con=ENGINE)

        if sheet.empty:
            print("Nenhum dado encontrado para o arquivo solicitado.")
            return
        
        sheet.rename(columns=colunas, inplace=True)

        if 'index' in sheet.columns:
            sheet.drop(columns=['index'], inplace=True)

        sheet.to_excel(file_path, index=False)

        print(f"Dados exportados com sucesso para o arquivo '{id_file}'.")
    except Exception as error:
        print(f"Erro ao extrair os dados do MySQL para o arquivo: {error}")


def normalize_text(text: str) -> str:
    """ Normaliza o texto removendo espaços extras e convertendo para minúsculas. """
    return ' '.join(text.strip().lower().split())


def extract_reference_urls(file_path: Path, cabecalho='Documento Referência') -> list[str]:
    """ Extrai as URLS de referência do arquivo Excel. """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.worksheets[0]

        reference_column_index = None

        for column in range(1, sheet.max_column + 1):
            cabecalho_value = sheet.cell(row=1, column=column).value
            cabecalho_norm = normalize_text(cabecalho_value)

            if cabecalho_norm in (normalize_text(cabecalho), 'documento referencia'):
                reference_column_index = column
                break

        if reference_column_index is None:
            print(f"Erro: Cabeçalho '{cabecalho}' não encontrado no arquivo.")
            return None

        urls = []

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=reference_column_index)
            url: str = None
            if cell.hyperlink is not None:
                url = getattr(cell.hyperlink, 'target', None) or str(cell.hyperlink)
            
            urls.append(url)

        return urls

    except Exception as error:
        print(f"Erro ao extrair URLs de referência: {error}")
        return []


if __name__ == '__main__':
    user_id = "806443c5-9271-464a-a1da-4581c7f766e4"
    id_file = "f9316b87-29f1-4a4b-a123-40b4517a3721.xlsx"
    filename = "Desafio Número 1_Projeto 6 - Exportado.xlsx"

    """to_mysql_inserir(id_file, filename, user_id)"""

    from_mysql_extrair(id_file)
